import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
from openpyxl import Workbook
import os
import sys


def get_app_folder():
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


APP_FOLDER = get_app_folder()
EXCEL_FILE = os.path.join(APP_FOLDER, "purchase_requests.xlsx")


def create_excel_file():
    folder = os.path.dirname(EXCEL_FILE)

    if folder and not os.path.exists(folder):
        os.makedirs(folder, exist_ok=True)

    if not os.path.exists(EXCEL_FILE):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Purchase Requests"

        sheet.append([
            "Artikelname",
            "Kategorie",
            "Menge",
            "Lieferant",
            "Einzelpreis",
            "Gesamtpreis",
            "Dringend",
            "Status"
        ])

        workbook.save(EXCEL_FILE)


def save_request():
    artikel = artikel_var.get().strip()
    kategorie = kategorie_var.get().strip()
    menge = menge_var.get().strip().replace(",", ".")
    lieferant = lieferant_var.get().strip()
    einzelpreis = einzelpreis_var.get().strip().replace(",", ".")
    dringend = dringend_var.get()
    status = status_var.get()

    if artikel == "" or menge == "" or einzelpreis == "":
        messagebox.showwarning(
            "Fehler",
            "Bitte Artikelname, Menge und Einzelpreis eingeben."
        )
        return

    try:
        menge_float = float(menge)
        einzelpreis_float = float(einzelpreis)
        gesamtpreis = menge_float * einzelpreis_float
    except ValueError:
        messagebox.showerror(
            "Fehler",
            "Menge und Einzelpreis müssen Zahlen sein."
        )
        return

    row_data = [
        artikel,
        kategorie,
        menge_float,
        lieferant,
        einzelpreis_float,
        gesamtpreis,
        "Ja" if dringend else "Nein",
        status
    ]

    try:
        create_excel_file()
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook.active
        sheet.append(row_data)
        workbook.save(EXCEL_FILE)
    except Exception as e:
        messagebox.showerror(
            "Speicherfehler",
            f"Datei konnte nicht gespeichert werden:\n{e}"
        )
        return

    tree.insert("", "end", values=row_data)

    status_label.config(text="Anfrage wurde gespeichert.")
    clear_fields()


def load_requests():
    try:
        create_excel_file()
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            tree.insert("", "end", values=row)

    except Exception as e:
        messagebox.showerror(
            "Ladefehler",
            f"Datei konnte nicht geladen werden:\n{e}"
        )


def clear_fields():
    artikel_var.set("")
    kategorie_var.set("")
    menge_var.set("")
    lieferant_var.set("")
    einzelpreis_var.set("")
    dringend_var.set(False)
    status_var.set("Offen")
    status_label.config(text="Felder wurden geleert.")


def exit_program():
    root.destroy()


create_excel_file()

root = tk.Tk()
root.title("Purchase Request Management Tool")
root.geometry("1000x650")

artikel_var = tk.StringVar()
kategorie_var = tk.StringVar()
menge_var = tk.StringVar()
lieferant_var = tk.StringVar()
einzelpreis_var = tk.StringVar()
dringend_var = tk.BooleanVar(value=False)
status_var = tk.StringVar(value="Offen")

menubar = tk.Menu(root)
file_menu = tk.Menu(menubar, tearoff=0)
file_menu.add_command(label="Beenden", command=exit_program)
menubar.add_cascade(label="Datei", menu=file_menu)
root.config(menu=menubar)

form_frame = tk.LabelFrame(root, text="Neue Einkaufsanfrage")
form_frame.pack(padx=10, pady=10, fill="x")

tk.Label(form_frame, text="Artikelname").grid(row=0, column=0, padx=5, pady=5)
tk.Entry(form_frame, textvariable=artikel_var).grid(row=0, column=1, padx=5, pady=5)

tk.Label(form_frame, text="Kategorie").grid(row=0, column=2, padx=5, pady=5)
kategorie_combo = ttk.Combobox(
    form_frame,
    textvariable=kategorie_var,
    values=["Rohmaterial", "Verpackung", "Ersatzteile", "Dienstleistung", "IT"],
    state="readonly"
)
kategorie_combo.grid(row=0, column=3, padx=5, pady=5)

tk.Label(form_frame, text="Menge").grid(row=1, column=0, padx=5, pady=5)
tk.Entry(form_frame, textvariable=menge_var).grid(row=1, column=1, padx=5, pady=5)

tk.Label(form_frame, text="Lieferant").grid(row=1, column=2, padx=5, pady=5)
tk.Entry(form_frame, textvariable=lieferant_var).grid(row=1, column=3, padx=5, pady=5)

tk.Label(form_frame, text="Einzelpreis").grid(row=2, column=0, padx=5, pady=5)
tk.Entry(form_frame, textvariable=einzelpreis_var).grid(row=2, column=1, padx=5, pady=5)

tk.Checkbutton(form_frame, text="Dringend", variable=dringend_var).grid(row=2, column=2, padx=5, pady=5)

status_frame = tk.Frame(form_frame)
status_frame.grid(row=3, column=0, columnspan=4, pady=5)

tk.Label(status_frame, text="Status:").pack(side="left")

tk.Radiobutton(status_frame, text="Offen", variable=status_var, value="Offen").pack(side="left")
tk.Radiobutton(status_frame, text="In Prüfung", variable=status_var, value="In Prüfung").pack(side="left")
tk.Radiobutton(status_frame, text="Genehmigt", variable=status_var, value="Genehmigt").pack(side="left")
tk.Radiobutton(status_frame, text="Abgelehnt", variable=status_var, value="Abgelehnt").pack(side="left")

button_frame = tk.Frame(root)
button_frame.pack(pady=10)

tk.Button(button_frame, text="Speichern", command=save_request).pack(side="left", padx=5)
tk.Button(button_frame, text="Felder leeren", command=clear_fields).pack(side="left", padx=5)
tk.Button(button_frame, text="Beenden", command=exit_program).pack(side="left", padx=5)

tree_frame = tk.LabelFrame(root, text="Gespeicherte Einkaufsanfragen")
tree_frame.pack(padx=10, pady=10, fill="both", expand=True)

columns = (
    "Artikelname",
    "Kategorie",
    "Menge",
    "Lieferant",
    "Einzelpreis",
    "Gesamtpreis",
    "Dringend",
    "Status"
)

tree = ttk.Treeview(tree_frame, columns=columns, show="headings")

for col in columns:
    tree.heading(col, text=col)
    tree.column(col, width=120)

tree.pack(fill="both", expand=True)

status_label = tk.Label(root, text="Bereit", relief="sunken", anchor="w")
status_label.pack(side="bottom", fill="x")

load_requests()

root.mainloop()